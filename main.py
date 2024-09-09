import pandas as pd
import streamlit as st
import io
import xlsxwriter
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import zipfile
# buffer to use for excel writer
buffer = io.BytesIO()


def template(group, col_to_group):
    # Percorso al file Excel caricato
    file_path = 'template.xlsx'

    # Leggere l'intero file Excel senza specificare un header
    df = pd.read_excel(file_path, header=None)
    ids = group[col_to_group]

    # Aggiungere le nuove righe con l'ID utente al DataFrame
    rows_to_add = []
    for row in ids:
        new_row = [f"{row}"] + [None] * (len(df.columns) - 1)
        rows_to_add.append(new_row)

    # Aggiungere le righe aggiuntive al DataFrame
    if rows_to_add:
        additional_df = pd.DataFrame(rows_to_add)
        df = pd.concat([df, additional_df], ignore_index=True)

    # Initialize a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write data starting from row 6
    for index, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)

    # Merge and center cells for the first 5 rows and columns A to E
    for i in range(1, 6):  # rows 1 to 5
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
        if i == 1:  # for the first row only
            ws.cell(row=i, column=1).font = Font(name='Calibri', bold=True)  # Setting bold font
        else:
            ws.cell(row=i, column=1).font = Font(name='Calibri', color='FF0000')  # Setting red font
     
    ws.cell(row=6, column=1).font = Font(name='Calibri', color='FF0000', bold=True)  # Setting bold and red font
    # Save the workbook
    return wb


#n_persone = 30
np.float = float
np.int = int

if __name__ == '__main__':

    st.set_page_config(page_title='XLSX Grouper', layout='wide')
    #IMAGE-UPLOAD FILE
    #b.title('XLSX Grouper')
    st.write('Dividere persone in diversi gruppi')
    #st.write(np.__version__)
    #st.write(pd.__version__)
    uploaded_file = st.file_uploader("Choose a file", type="xlsx")
    #DASHBOARD
    if uploaded_file is not None:
    # To read file as bytes:
        st.write(f'You selected {uploaded_file.name}')
        df = pd.read_excel(uploaded_file).reset_index(drop=True)
        columns = df.columns
        col_to_group = st.selectbox("Colonna da usare per fare ordinamento:", columns )
        n_sessione = st.number_input("numero di gruppi", min_value=1, format="%d")
        m = st.number_input("numero di persone massimo per gruppo", min_value=1, format="%d")
        if st.button("RUN"):

            group = ['Utente - Nome utente', 'Utente - Cognome utente', 'Utente - Luogo ID']
            df[group] = df[group].apply(lambda x: x.str.upper())
            gb = df.groupby(group).agg(list).reset_index()
            gb = gb.sort_values(by=col_to_group)
            gb["session"] =  list(range(1, len(gb) + 1))
            gb["session"] =  gb["session"].apply(lambda x: x % n_sessione)
            gb["session"] =  gb["session"]+1
            gb = gb.apply(lambda x: x if len(x) <= m else x.head(m))
            df= gb.explode("Utente - ID utente")
            df = df.sort_values(by=["session"]+group)
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
               # Close the Pandas Excel writer and output the Excel file to the buffer
               # writer.save()

            download2 = st.download_button(
            label="Download data as Excel",
            data=buffer,
            file_name='Executed.xlsx',
            mime='application/vnd.ms-excel'
            )
            gb2 = df.groupby("session")
            # Creazione di un buffer di byte in memoria per lo zip
            buffer = io.BytesIO()
            zipf = zipfile.ZipFile(buffer, 'w')

            # Ciclo for per chiamare template() e aggiungere i file Excel allo zip
            for group in gb2.groups:
                # Chiamata alla funzione template per ottenere l'oggetto Workbook
                wb = template(gb2.get_group(group), col_to_group)  # 'output' può essere None nel contesto di questo esempio

                # Salvataggio del file Excel generato nello zip
                excel_file_name = f'file_{group}.xlsx'
                wb.save(excel_file_name)
                zipf.write(excel_file_name)

                # Chiudi il Workbook dopo averlo salvato nello zip
                wb.close()

            # Chiudi lo zip una volta completato
            zipf.close()

            # Resetta il buffer per consentire di leggere il suo contenuto
            buffer.seek(0)
            # Creazione di un pulsante di download in Streamlit per scaricare lo zip
            st.download_button(
                label='Download ZIP',
                data=buffer,
                file_name='excel_files.zip',
                mime='application/zip'
            )
